import os
import csv
import io
import re
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pymongo import MongoClient
from datetime import datetime, timezone, timedelta
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

from registry import VARIABLES, SYSTEMS, ALERT_RULES, GROUPING_UNITS, get_nested

# Spanish column headers used in every export
_HEADERS_ES  = ["Fecha y Hora", "Promedio", "Mínimo", "Máximo", "Cantidad de registros"]
_ROW_KEYS    = ["timestamp",    "avg",      "min",    "max",    "count"]


def _safe_name(text: str) -> str:
    """Strip characters unsafe for filenames."""
    return re.sub(r"[^\w\-]", "_", text)


def _fmt_ts(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val) if val is not None else ""


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def _csv_single(var_info: dict, system_name: str, data: list) -> io.StringIO:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"Sistema: {system_name}"])
    w.writerow([f"Variable: {var_info['name']} ({var_info['measurementUnit']})"])
    w.writerow([])
    w.writerow(_HEADERS_ES)
    for row in data:
        w.writerow([
            _fmt_ts(row.get("timestamp")),
            row.get("avg", ""), row.get("min", ""),
            row.get("max", ""), row.get("count", ""),
        ])
    buf.seek(0)
    return buf


def _csv_all(system_name: str, variables_data: dict) -> io.StringIO:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"Sistema: {system_name}"])
    w.writerow([])
    w.writerow(["Variable", "Unidad"] + _HEADERS_ES)
    for var_data in variables_data.values():
        vi = var_data["variable"]
        for row in var_data["data"]:
            w.writerow([
                vi["name"], vi["measurementUnit"],
                _fmt_ts(row.get("timestamp")),
                row.get("avg", ""), row.get("min", ""),
                row.get("max", ""), row.get("count", ""),
            ])
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, row_num: int, ncols: int):
    fill = PatternFill("solid", fgColor="2E7D32")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _excel_single(var_info: dict, system_name: str, data: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = var_info["name"][:31]

    ws.append([f"Sistema: {system_name}"])
    ws.append([f"Variable: {var_info['name']} ({var_info['measurementUnit']})"])
    ws.append([])
    ws.append(_HEADERS_ES)
    _style_header_row(ws, 4, len(_HEADERS_ES))

    for row in data:
        ws.append([
            _fmt_ts(row.get("timestamp")),
            row.get("avg", ""), row.get("min", ""),
            row.get("max", ""), row.get("count", ""),
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _excel_all(system_name: str, variables_data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for var_data in variables_data.values():
        vi  = var_data["variable"]
        ws  = wb.create_sheet(title=vi["name"][:31])
        ws.append([f"Sistema: {system_name}"])
        ws.append([f"Variable: {vi['name']} ({vi['measurementUnit']})"])
        ws.append([])
        ws.append(_HEADERS_ES)
        _style_header_row(ws, 4, len(_HEADERS_ES))

        for row in var_data["data"]:
            ws.append([
                _fmt_ts(row.get("timestamp")),
                row.get("avg", ""), row.get("min", ""),
                row.get("max", ""), row.get("count", ""),
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

router = APIRouter()
client = MongoClient(os.getenv("MONGO_URI", "mongodb://localhost:27017"))
col = client[os.getenv("MONGO_DB", "terrafy")]["readings"]


def _utc(dt: datetime) -> datetime:
    """Ensure datetime is UTC-aware."""
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


# ---------------------------------------------------------------------------
# Registry info
# ---------------------------------------------------------------------------

@router.get("/variables")
def list_variables():
    """Return all registered agronomic variables."""
    return list(VARIABLES.values())


@router.get("/systems")
def list_systems():
    """Return all registered agronomic systems."""
    return list(SYSTEMS.values())


# ---------------------------------------------------------------------------
# Latest snapshot (kept for backward compat)
# ---------------------------------------------------------------------------

@router.get("/latest")
def get_latest(system_id: Optional[str] = None):
    query = {"system_id": system_id} if system_id else {}
    doc = col.find_one(query, sort=[("timestamp", -1)])
    if doc:
        doc.pop("_id")
    return doc or {}


# ---------------------------------------------------------------------------
# Historic data — single variable for a system
# ---------------------------------------------------------------------------

@router.get("/history")
def get_history(
    system_id: str,
    variable_id: str,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    grouping: str = Query("hours", enum=["minutes", "hours", "days", "weeks"]),
    formato: Optional[str] = Query(None, enum=["csv", "excel"]),
):
    """
    Return time-grouped statistics for one agronomic variable on one system.

    - system_id:    e.g. 1, 2, 3
    - variable_id:  e.g. 1 (Temperatura), 4 (pH), 5 (EC) …
    - start_date / end_date: ISO-8601 (defaults to last 24 h)
    - grouping:     minutes | hours | days | weeks
    - formato:      csv | excel  (omit for JSON)
    """
    if variable_id not in VARIABLES:
        raise HTTPException(
            400,
            detail=f"Unknown variable '{variable_id}'. Valid: {list(VARIABLES.keys())}",
        )
    if system_id not in SYSTEMS:
        raise HTTPException(
            400,
            detail=f"Unknown system '{system_id}'. Valid: {list(SYSTEMS.keys())}",
        )
    if formato == "excel" and not EXCEL_OK:
        raise HTTPException(503, detail="openpyxl not installed — Excel export unavailable.")

    var  = VARIABLES[variable_id]
    sys_ = SYSTEMS[system_id]
    now  = datetime.now(timezone.utc)
    start = _utc(start_date) if start_date else now - timedelta(hours=24)
    end   = _utc(end_date)   if end_date   else now
    unit  = GROUPING_UNITS[grouping]

    pipeline = [
        {"$match": {"system_id": system_id, "timestamp": {"$gte": start, "$lte": end}}},
        {"$addFields": {
            "bucket": {"$dateTrunc": {"date": "$timestamp", "unit": unit}}
        }},
        {"$group": {
            "_id":   "$bucket",
            "avg":   {"$avg": f"${var['field']}"},
            "min":   {"$min": f"${var['field']}"},
            "max":   {"$max": f"${var['field']}"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
        {"$project": {
            "_id": 0,
            "timestamp": "$_id",
            "avg":   {"$round": ["$avg",   4]},
            "min":   {"$round": ["$min",   4]},
            "max":   {"$round": ["$max",   4]},
            "count": 1,
        }},
    ]

    data = list(col.aggregate(pipeline))

    if formato == "csv":
        fname = f"historico_{_safe_name(sys_['name'])}_{_safe_name(var['name'])}.csv"
        buf   = _csv_single(var, sys_["name"], data)
        return StreamingResponse(
            buf,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    if formato == "excel":
        fname = f"historico_{_safe_name(sys_['name'])}_{_safe_name(var['name'])}.xlsx"
        buf   = _excel_single(var, sys_["name"], data)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    return {
        "system_id":   system_id,
        "system_name": sys_["name"],
        "variable":    var,
        "grouping":    grouping,
        "start_date":  start,
        "end_date":    end,
        "data":        data,
    }


# ---------------------------------------------------------------------------
# Historic data — all variables for a system
# ---------------------------------------------------------------------------

@router.get("/history/{system_id}")
def get_history_by_system(
    system_id: str,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    grouping: str = Query("hours", enum=["minutes", "hours", "days", "weeks"]),
    formato: Optional[str] = Query(None, enum=["csv", "excel"]),
):
    """
    Return time-grouped statistics for every agronomic variable of a system.

    - grouping: minutes | hours | days | weeks
    - formato:  csv | excel  (omit for JSON)
    """
    if system_id not in SYSTEMS:
        raise HTTPException(400, detail=f"Unknown system '{system_id}'")
    if formato == "excel" and not EXCEL_OK:
        raise HTTPException(503, detail="openpyxl not installed — Excel export unavailable.")

    sys_  = SYSTEMS[system_id]
    now   = datetime.now(timezone.utc)
    start = _utc(start_date) if start_date else now - timedelta(hours=24)
    end   = _utc(end_date)   if end_date   else now
    unit  = GROUPING_UNITS[grouping]

    match = {"system_id": system_id, "timestamp": {"$gte": start, "$lte": end}}

    variables_data = {}
    for var_name, var in VARIABLES.items():
        pipeline = [
            {"$match": match},
            {"$addFields": {
                "bucket": {"$dateTrunc": {"date": "$timestamp", "unit": unit}}
            }},
            {"$group": {
                "_id":   "$bucket",
                "avg":   {"$avg": f"${var['field']}"},
                "min":   {"$min": f"${var['field']}"},
                "max":   {"$max": f"${var['field']}"},
                "count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
            {"$project": {
                "_id": 0,
                "timestamp": "$_id",
                "avg":   {"$round": ["$avg",   4]},
                "min":   {"$round": ["$min",   4]},
                "max":   {"$round": ["$max",   4]},
                "count": 1,
            }},
        ]
        variables_data[var_name] = {
            "variable": var,
            "data":     list(col.aggregate(pipeline)),
        }

    if formato == "csv":
        fname = f"historico_{_safe_name(sys_['name'])}.csv"
        buf   = _csv_all(sys_["name"], variables_data)
        return StreamingResponse(
            buf,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    if formato == "excel":
        fname = f"historico_{_safe_name(sys_['name'])}.xlsx"
        buf   = _excel_all(sys_["name"], variables_data)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    return {
        "system_id":   system_id,
        "system_name": sys_["name"],
        "grouping":    grouping,
        "start_date":  start,
        "end_date":    end,
        "variables":   variables_data,
    }


# ---------------------------------------------------------------------------
# Alerts
# ---------------------------------------------------------------------------

@router.get("/alerts")
def get_alerts(
    system_id: str,
    variable_id: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
):
    """
    Return threshold violations for a system, optionally filtered to one variable.

    Alertable variable IDs: 3 (VPD), 4 (pH), 5 (EC), 7 (Nitrógeno).
    """
    if system_id not in SYSTEMS:
        raise HTTPException(400, detail=f"Unknown system '{system_id}'")
    if variable_id and variable_id not in ALERT_RULES:
        raise HTTPException(
            400,
            detail=f"Variable '{variable_id}' has no alert rules. Alertable IDs: {list(ALERT_RULES.keys())}",
        )

    now   = datetime.now(timezone.utc)
    start = _utc(start_date) if start_date else now - timedelta(hours=24)
    end   = _utc(end_date)   if end_date   else now

    docs = list(col.find(
        {"system_id": system_id, "timestamp": {"$gte": start, "$lte": end}},
        {"_id": 0},
    ))

    rules = (
        {variable_id: ALERT_RULES[variable_id]}
        if variable_id
        else ALERT_RULES
    )

    alerts = []
    for doc in docs:
        for vid, rule in rules.items():
            val = get_nested(doc, rule["field"])
            if val is None:
                continue
            direction = None
            if rule["min"] is not None and val < rule["min"]:
                direction = "below_minimum"
            elif rule["max"] is not None and val > rule["max"]:
                direction = "above_maximum"
            if direction:
                alerts.append({
                    "type":          f"variable_{vid}_out_of_range",
                    "variable_id":   vid,
                    "variable_name": VARIABLES[vid]["name"],
                    "value":         round(val, 4),
                    "unit":          rule["unit"],
                    "direction":     direction,
                    "system_id":     system_id,
                    "timestamp":     doc["timestamp"],
                })

    return {
        "system_id":   system_id,
        "system_name": SYSTEMS[system_id]["name"],
        "start_date":  start,
        "end_date":    end,
        "alert_count": len(alerts),
        "alerts":      alerts[-50:],
    }
